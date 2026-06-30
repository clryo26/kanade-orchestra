from io import BytesIO
from pathlib import Path
import sys
import types
import importlib.util
from fastapi import APIRouter
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
BACKEND = ROOT / "src" / "backend"
sys.path.insert(0, str(BACKEND))

# main.py 末尾の auth_api import を回避するためダミーを差し込む
sys.modules["auth_api"] = types.SimpleNamespace(router=APIRouter())

spec = importlib.util.spec_from_file_location("backend_main", BACKEND / "main.py")
mod = importlib.util.module_from_spec(spec)
spec.loader.exec_module(mod)

performance = {
    "id": 1,
    "title": "テスト演奏会",
    "date": "2026-07-20",
    "pieces": [
        {"title": "アンパンマン", "duration": "7"},
        {"title": "ナウシカ", "duration": "6"},
    ],
}
info = {
    "performance_id": "1",
    "timeline_rows": [
        {
            "sort_order": 1,
            "start_time": "09:00",
            "end_time": "09:40",
            "duration_minutes": "40",
            "section": "準備",
            "content": "ステージ準備",
            "mc": "岡内",
            "reception": "森光さん",
            "setting": "摂津/渡辺/古賀",
            "note": "",
        },
        {
            "sort_order": 2,
            "start_time": "09:40",
            "end_time": "09:47",
            "duration_minutes": "7",
            "section": "第一部",
            "content": "アンパンマン",
            "mc": "",
            "reception": "",
            "setting": "",
            "note": "",
        },
    ],
    "assignments_rows": [
        {"role": "受付", "members": "森光さん"},
        {"role": "MC", "members": "岡内"},
        {"role": "セッティング", "members": "摂津/渡辺/古賀"},
    ],
}

binary = mod.build_timetable_workbook_bytes(performance, info)
out = ROOT / "sample" / "本番タイムテーブル_自動出力確認.xlsx"
out.write_bytes(binary)

wb = load_workbook(BytesIO(binary))
ws = wb.active
print(f"OUTPUT={out.as_posix()}")
print("B4", ws["B4"].value)
for row in (4, 5, 6):
    print(
        row,
        ws[f"C{row}"].value,
        ws[f"D{row}"].value,
        ws[f"E{row}"].value,
        ws[f"F{row}"].value,
        ws[f"G{row}"].value,
        ws[f"H{row}"].value,
        ws[f"I{row}"].value,
        ws[f"J{row}"].value,
        ws[f"K{row}"].value,
    )
