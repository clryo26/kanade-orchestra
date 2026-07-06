from __future__ import annotations
# mypy: disable-error-code="no-redef,misc"

import io  # noqa: F401
import os
import re  # noqa: F401
from functools import partial
from datetime import date as date  # noqa: F401
from datetime import datetime as datetime  # noqa: F401
from datetime import time as time  # noqa: F401
from decimal import Decimal as Decimal  # noqa: F401
from pathlib import Path as Path  # noqa: F401
from typing import Any
from urllib.parse import quote as quote  # noqa: F401

from fastapi import HTTPException
from fastapi import Request as Request  # noqa: F401
from fastapi.responses import Response as Response  # noqa: F401

try:
    from .models.schemas import (
        ExtraUpsertRequest as ExtraUpsertRequest,  # noqa: F401
        Member as Member,  # noqa: F401
    )
    from .models.schemas import Announcement as Announcement  # noqa: F401
    from .models.schemas import EventAdjustment as EventAdjustment  # noqa: F401
    from .models.schemas import MemberPasswordSetupRequest as MemberPasswordSetupRequest  # noqa: F401
    from .models.schemas import Performance as Performance  # noqa: F401
    from .models.schemas import PortalLoginRequest as PortalLoginRequest  # noqa: F401
    from .models.schemas import RecordingDeleteRequest as RecordingDeleteRequest  # noqa: F401
    from .models.schemas import Schedule as Schedule  # noqa: F401
    from .models.schemas import SheetBulkPartUpdateRequest as SheetBulkPartUpdateRequest  # noqa: F401
    from .models.schemas import SheetDeleteRequest as SheetDeleteRequest  # noqa: F401
    from .models.schemas import SheetPartUpdateRequest as SheetPartUpdateRequest  # noqa: F401
    from .services.security_service import hash_password as hash_password  # noqa: F401
    from .services.security_service import is_hashed_password as is_hashed_password  # noqa: F401
    from .services.security_service import is_password_placeholder as is_password_placeholder  # noqa: F401
    from .services.security_service import verify_password as verify_password  # noqa: F401
except ImportError:  # pragma: no cover - allows running main.py directly.
    from models.schemas import (
        ExtraUpsertRequest as ExtraUpsertRequest,  # noqa: F401
        Member as Member,  # noqa: F401
    )
    from models.schemas import Announcement as Announcement  # noqa: F401
    from models.schemas import EventAdjustment as EventAdjustment  # noqa: F401
    from models.schemas import MemberPasswordSetupRequest as MemberPasswordSetupRequest  # noqa: F401
    from models.schemas import Performance as Performance  # noqa: F401
    from models.schemas import PortalLoginRequest as PortalLoginRequest  # noqa: F401
    from models.schemas import RecordingDeleteRequest as RecordingDeleteRequest  # noqa: F401
    from models.schemas import Schedule as Schedule  # noqa: F401
    from models.schemas import SheetBulkPartUpdateRequest as SheetBulkPartUpdateRequest  # noqa: F401
    from models.schemas import SheetDeleteRequest as SheetDeleteRequest  # noqa: F401
    from models.schemas import SheetPartUpdateRequest as SheetPartUpdateRequest  # noqa: F401
    from services.security_service import hash_password as hash_password  # noqa: F401
    from services.security_service import is_hashed_password as is_hashed_password  # noqa: F401
    from services.security_service import is_password_placeholder as is_password_placeholder  # noqa: F401
    from services.security_service import verify_password as verify_password  # noqa: F401

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None

try:
    import psycopg
    from psycopg import sql as psql
    from psycopg.types.json import Jsonb
except ImportError:  # pragma: no cover
    psycopg = None
    psql = None
    Jsonb = None

try:
    from .drive_storage import (
        get_storage_bucket as get_storage_bucket,  # noqa: F401
        storage_enabled as storage_enabled,  # noqa: F401
        upload_file_to_drive,  # noqa: F401
    )
except ImportError as exc:  # pragma: no cover - allows running tests without optional Google deps.
    # If the optional google-cloud-storage dependency is missing in a local test
    # environment, keep the application importable and report storage as disabled.
    # When the file is executed directly, fall back to the non-package import.
    if getattr(exc, "name", "") and str(exc.name).startswith("google"):
        def get_storage_bucket():
            return None

        def storage_enabled() -> bool:
            return False

        def upload_file_to_drive(*args, **kwargs):  # noqa: ANN001
            raise HTTPException(status_code=500, detail="Google Cloud Storage dependency is not installed")
    else:
        from drive_storage import (
            get_storage_bucket as get_storage_bucket,  # noqa: F401
            storage_enabled as storage_enabled,  # noqa: F401
            upload_file_to_drive,  # noqa: F401
        )

try:
    from .auth_helpers import (
        member_login_names as member_login_names,  # noqa: F401
    )
    from .auth_helpers import member_access_expired as member_access_expired  # noqa: F401
    from .auth_helpers import member_display_name as member_display_name  # noqa: F401
except ImportError:  # pragma: no cover - allows running main.py directly.
    from auth_helpers import (
        member_login_names as member_login_names,  # noqa: F401
    )
    from auth_helpers import member_access_expired as member_access_expired  # noqa: F401
    from auth_helpers import member_display_name as member_display_name  # noqa: F401

try:
    from .core.db_schema import (
        DB_CHILD_COLUMNS as DB_CHILD_COLUMNS,  # noqa: F401
        DB_COLLECTION_COLUMNS as DB_COLLECTION_COLUMNS,  # noqa: F401
        DB_INT_COLUMNS as DB_INT_COLUMNS,  # noqa: F401
        DB_JSON_COLUMNS as DB_JSON_COLUMNS,  # noqa: F401
        DB_TIMESTAMP_COLUMNS as DB_TIMESTAMP_COLUMNS,  # noqa: F401
        DB_WRITABLE_COLLECTIONS,
        JSON_COLLECTION_TABLES,
        PORTAL_DB_TABLES as PORTAL_DB_TABLES,  # noqa: F401
    )
except ImportError:  # pragma: no cover - allows running main.py directly.
    from core.db_schema import (
        DB_CHILD_COLUMNS as DB_CHILD_COLUMNS,  # noqa: F401
        DB_COLLECTION_COLUMNS as DB_COLLECTION_COLUMNS,  # noqa: F401
        DB_INT_COLUMNS as DB_INT_COLUMNS,  # noqa: F401
        DB_JSON_COLUMNS as DB_JSON_COLUMNS,  # noqa: F401
        DB_TIMESTAMP_COLUMNS as DB_TIMESTAMP_COLUMNS,  # noqa: F401
        DB_WRITABLE_COLLECTIONS,
        JSON_COLLECTION_TABLES,
        PORTAL_DB_TABLES as PORTAL_DB_TABLES,  # noqa: F401
    )

from .core.compat_runtime import (
    app,
    effective_local_json_fallback_enabled,
    logger,
    memory_cache_instance,
)
from .core.compat_helpers import (
    load_json_data_compat,
    save_json_data_compat,
    seed_cloud_data_from_local_compat,
    seed_connection_settings_from_legacy_env_compat,
)
from .core.compat_data import (
    check_etag as compat_check_etag,
    find_item as compat_find_item,
    list_auth_devices as compat_list_auth_devices,
    load_json_data as compat_load_json_data,
    save_json_data as compat_save_json_data,
    seed_cloud_data_from_local as compat_seed_cloud_data_from_local,
    seed_connection_settings_from_legacy_env as compat_seed_connection_settings_from_legacy_env,
)
from .core.compat_recordings import (
    cloud_recording_metadata as compat_cloud_recording_metadata,
    forget_drive_file as compat_forget_drive_file,
    local_recording_metadata as compat_local_recording_metadata,
    local_recording_path as compat_local_recording_path,
    recording_file_bytes as compat_recording_file_bytes,
    recording_metadata_map as compat_recording_metadata_map,
    recording_payload as compat_recording_payload,
    remember_drive_file as compat_remember_drive_file,
    remember_recording_duration as compat_remember_recording_duration,
)
from .core.app_lifecycle import (
    db_expected as lifecycle_db_expected,
    ensure_db_expected_is_ready as lifecycle_ensure_db_expected_is_ready,
    local_json_fallback_enabled as lifecycle_local_json_fallback_enabled,
    run_db_startup_self_check as lifecycle_run_db_startup_self_check,
)
from .core.compat_startup import (
    has_connection_setting as startup_has_connection_setting,
    legacy_connection_setting_from_env as startup_legacy_connection_setting_from_env,
    list_auth_devices as startup_list_auth_devices,
    seed_cloud_data_from_local as startup_seed_cloud_data_from_local,
    seed_connection_settings_from_legacy_env as startup_seed_connection_settings_from_legacy_env,
)
from .core.runtime_paths import (
    CONVERTED_DIR as CONVERTED_DIR,  # noqa: F401
    DATA_DIR,
    DRIVE_STAGING_DIR as DRIVE_STAGING_DIR,  # noqa: F401
    JSON_DATA_NAMES,
    SHEET_DIR as SHEET_DIR,  # noqa: F401
    STARTUP_PRELOAD_COLLECTIONS,
    STATIC_DIR as STATIC_DIR,  # noqa: F401
    TIMETABLE_TEMPLATE_PATH,
    UPLOAD_DIR as UPLOAD_DIR,  # noqa: F401
)
from .core.app_core_exports import APP_CORE_EXPORTS
from .core.authorization import require_admin_device as _require_admin_device
from .core.authorization import require_device as _require_device
from .core.authorization import require_system_admin_device as _require_system_admin_device
from .core.database import assert_db_ready as _assert_db_ready
from .core.database import db_connection_string as db_connection_string
from .core.database import db_configured as _db_configured
from .core.database import mask_db_value as mask_db_value
from .services.auth_service import (
    device_auth_record as device_auth_record,
    prepare_member_payload as prepare_member_payload,
    public_member_list as public_member_list,
    public_member_payload as public_member_payload,
    require_recording_manager_device as require_recording_manager_device,
    require_sheet_manager_device as require_sheet_manager_device,
)
from .services.file_service import (
    ensure_audio_file as ensure_audio_file,
    ensure_pdf_file as ensure_pdf_file,
    format_duration as format_duration,
    safe_segment as safe_segment,
    safe_upload_name as safe_upload_name,
    save_upload_to_path as save_upload_to_path,
)
from .services.audio_processing_service import convert_path_to_mp3 as _convert_path_to_mp3
from .services.audio_processing_service import get_audio_duration_seconds as _get_audio_duration_seconds
from .services.extra_collection_helpers import EXTRA_COLLECTIONS
from .services.extra_collection_helpers import assert_extra_collection_permission as assert_extra_collection_permission
from .services.extra_collection_helpers import parse_extra_upsert_request as parse_extra_upsert_request
from .services.extra_collection_helpers import read_json_body as read_json_body
from .services.recording_asset_service import cloud_recording_metadata as _cloud_recording_metadata
from .services.recording_asset_service import forget_drive_file as _forget_drive_file
from .services.recording_asset_service import local_recording_metadata as _local_recording_metadata
from .services.recording_asset_service import local_recording_path as _local_recording_path
from .services.recording_asset_service import recording_file_bytes as _recording_file_bytes
from .services.recording_asset_service import recording_metadata_map as _recording_metadata_map
from .services.recording_asset_service import recording_payload as _recording_payload
from .services.recording_asset_service import remember_drive_file as _remember_drive_file
from .services.recording_asset_service import remember_recording_duration as _remember_recording_duration
from .services.json_collection_service import data_file as _data_file
from .services.json_collection_service import load_local_json_data as _load_local_json_data
from .services.meta_service import cloud_run_revision as cloud_run_revision
from .repositories.db_json_repository import load_json_data as db_load_json_data
from .repositories.db_json_repository import replace_collection as db_replace_collection
from .repositories.db_json_repository import upsert_auth_device as db_upsert_auth_device  # noqa: F401
from .repositories.db_json_repository import delete_auth_device as db_delete_auth_device  # noqa: F401
from .repositories.db_json_repository import load_generic_json_collection as db_load_generic_json_collection  # noqa: F401
from .repositories.db_json_repository import save_generic_json_collection as db_save_generic_json_collection  # noqa: F401
from .repositories.db_row_repository import db_child_rows_for_collection as db_child_rows_for_collection
from .repositories.db_row_repository import db_collection_rows_for_save as db_collection_rows_for_save
from .repositories.db_row_repository import db_delete_collection_children as db_delete_collection_children
from .repositories.db_row_repository import db_fetch_all as db_fetch_all
from .repositories.db_row_repository import db_fill_missing_ids as db_fill_missing_ids
from .repositories.db_row_repository import db_insert_rows as db_insert_rows
from .repositories.db_row_repository import db_item_value as db_item_value
from .repositories.db_row_repository import db_json_value as db_json_value
from .repositories.db_row_repository import db_next_id as db_next_id
from .repositories.db_row_repository import db_row_to_json as db_row_to_json
from .repositories.db_row_repository import db_row_tuple as db_row_tuple
from .repositories.db_row_repository import db_upsert_rows as db_upsert_rows
from .repositories.db_row_repository import db_write_value as db_write_value
from .repositories.db_row_repository import parse_db_date as parse_db_date
from .repositories.db_row_repository import parse_db_month as parse_db_month
from .repositories.db_row_repository import parse_db_time as parse_db_time
from .repositories.db_row_repository import parse_db_timestamp as parse_db_timestamp
from .core.db_runtime import db_data_enabled as _db_data_enabled
from .core.db_runtime import ensure_db_schema_compatibility as ensure_db_schema_compatibility
from .core.db_runtime import env_flag_enabled as env_flag_enabled
from .utils.concurrency import ensure_expected_updated_at as ensure_expected_updated_at
from .utils.datetime_utils import next_updated_at as next_updated_at
from .services.sheet_asset_service import (
    delete_sheet_file as delete_sheet_file,
    local_sheet_path as local_sheet_path,
    sheet_file_bytes as sheet_file_bytes,
    sheet_metadata as sheet_metadata,
    sheet_payload as _sheet_payload,
    unique_zip_name as unique_zip_name,
)
from .services.timetable_payload_helpers import (
    add_minutes_to_clock_text as add_minutes_to_clock_text,
    candidate_sort_key as candidate_sort_key,
    choose_assignment_value as choose_assignment_value,
    clock_to_time as clock_to_time,
    compact_assignment_text as compact_assignment_text,
    excel_row_count_from_template as excel_row_count_from_template,
    excel_safe_filename as excel_safe_filename,
    infer_duration_from_content as infer_duration_from_content,
    normalize_bool_text as normalize_bool_text,
    normalize_clock_text as normalize_clock_text,
    normalize_extra_for_collection as normalize_extra_for_collection,
    parse_assignment_rows as parse_assignment_rows,
    parse_timeline_text_rows as parse_timeline_text_rows,
    performance_piece_labels as performance_piece_labels,
    set_sheet_value_if_writable as set_sheet_value_if_writable,
    validate_connection_settings_payload as validate_connection_settings_payload,
    validate_date_adjustment_payload as validate_date_adjustment_payload,
    validate_date_adjustment_response_payload as validate_date_adjustment_response_payload,
)
from .services.timetable_payload_helpers import build_timetable_workbook_bytes as _build_timetable_workbook_bytes
from .services.timetable_payload_helpers import normalized_timeline_rows as normalized_timeline_rows
from .services.timetable_payload_helpers import performance_day_info_for_performance as _performance_day_info_for_performance
from .utils.collection_utils import check_etag as _check_etag, find_item as _find_item, next_id as _next_id
from .utils.serialization import fk_int as fk_int
from .utils.serialization import model_dump as model_dump
_memory_cache = memory_cache_instance()


def performance_day_info_for_performance(performance_id):
    return _performance_day_info_for_performance(performance_id, load_json_data)


build_timetable_workbook_bytes = partial(_build_timetable_workbook_bytes, load_workbook_func=load_workbook, template_path=TIMETABLE_TEMPLATE_PATH)

_PBKDF2_ALGO = "sha256"
_PBKDF2_ITERATIONS = 260000  # OWASP 2023謗ｨ螂ｨ蛟､


require_device = partial(_require_device, device_auth_record=device_auth_record)
require_admin_device = partial(_require_admin_device, device_auth_record=device_auth_record)
require_system_admin_device = partial(_require_system_admin_device, device_auth_record=device_auth_record)


# Revision endpoint moved to src/backend/routers/meta.py.


assert_db_ready = partial(_assert_db_ready, psycopg_module=psycopg, psql_module=psql)
db_configured = partial(_db_configured, psycopg_module=psycopg, psql_module=psql)
db_data_enabled = partial(_db_data_enabled, psycopg=psycopg, psql=psql)


def local_json_fallback_enabled() -> bool:
    return lifecycle_local_json_fallback_enabled(env_flag_enabled=env_flag_enabled)


db_expected = partial(
    lifecycle_db_expected,
    local_json_fallback_enabled=local_json_fallback_enabled,
    env_flag_enabled=env_flag_enabled,
)
ensure_db_expected_is_ready = partial(
    lifecycle_ensure_db_expected_is_ready,
    db_expected=db_expected,
    db_data_enabled=db_data_enabled,
)
def run_db_startup_self_check() -> None:
    lifecycle_run_db_startup_self_check(
        assert_db_ready=assert_db_ready,
        db_connection_string=db_connection_string,
        ensure_db_schema_compatibility=ensure_db_schema_compatibility,
        psycopg=psycopg,
        db_expected=db_expected,
        ensure_db_expected_is_ready=ensure_db_expected_is_ready,
    )


data_file = partial(_data_file, data_dir=DATA_DIR)
load_local_json_data = partial(_load_local_json_data, data_dir=DATA_DIR, logger=logger)


def load_json_data(name: str) -> list[dict[str, Any]]:
    return compat_load_json_data(
        name,
        load_json_data_compat=load_json_data_compat,
        cache=_memory_cache,
        effective_local_json_fallback_enabled=effective_local_json_fallback_enabled,
        db_data_enabled=db_data_enabled,
        db_expected=db_expected,
        local_json_fallback_enabled=local_json_fallback_enabled,
        ensure_db_expected_is_ready=ensure_db_expected_is_ready,
        db_load_json_data=db_load_json_data,
        db_load_generic_json_collection=db_load_generic_json_collection,
        json_collection_tables=JSON_COLLECTION_TABLES,
        json_data_names=JSON_DATA_NAMES,
        extra_collections=EXTRA_COLLECTIONS,
        data_dir=DATA_DIR,
        logger=logger,
    )


def save_json_data(name: str, data: list[dict[str, Any]]) -> None:
    compat_save_json_data(
        name,
        data,
        save_json_data_compat=save_json_data_compat,
        cache=_memory_cache,
        effective_local_json_fallback_enabled=effective_local_json_fallback_enabled,
        db_data_enabled=db_data_enabled,
        db_expected=db_expected,
        local_json_fallback_enabled=local_json_fallback_enabled,
        ensure_db_expected_is_ready=ensure_db_expected_is_ready,
        db_replace_collection=db_replace_collection,
        db_save_generic_json_collection=db_save_generic_json_collection,
        db_writable_collections=DB_WRITABLE_COLLECTIONS,
        json_data_names=JSON_DATA_NAMES,
        extra_collections=EXTRA_COLLECTIONS,
        data_dir=DATA_DIR,
    )


has_connection_setting = startup_has_connection_setting
legacy_connection_setting_from_env = startup_legacy_connection_setting_from_env


seed_connection_settings_from_legacy_env = partial(
    compat_seed_connection_settings_from_legacy_env,
    compat_func=partial(
        seed_connection_settings_from_legacy_env_compat,
        startup_seed_connection_settings_from_legacy_env=startup_seed_connection_settings_from_legacy_env,
    ),
    load_json_data=load_json_data,
    save_json_data=save_json_data,
    next_id=lambda items: _next_id(items),
    logger=logger,
)
async def seed_cloud_data_from_local() -> None:
    await compat_seed_cloud_data_from_local(
        compat_func=partial(
            seed_cloud_data_from_local_compat,
            startup_seed_cloud_data_from_local=startup_seed_cloud_data_from_local,
        ),
        seed_connection_settings_from_legacy_env=seed_connection_settings_from_legacy_env,
        load_json_data=load_json_data,
        save_json_data=save_json_data,
        next_id=lambda items: _next_id(items),
        logger=logger,
        startup_preload_collections=STARTUP_PRELOAD_COLLECTIONS,
        db_expected=db_expected,
    )


next_id = _next_id
find_item = partial(compat_find_item, compat_func=_find_item, cache=_memory_cache, cache_names=JSON_DATA_NAMES)
check_etag = partial(compat_check_etag, compat_func=_check_etag, cache=_memory_cache)


# Bootstrap helpers moved to src/backend/services/bootstrap_service.py.

list_auth_devices = partial(compat_list_auth_devices, compat_func=startup_list_auth_devices, load_json_data=load_json_data)
# Access log and bootstrap endpoints moved to dedicated routers.


recording_metadata_map = partial(compat_recording_metadata_map, compat_func=_recording_metadata_map, load_json_data=load_json_data)
local_recording_metadata = partial(
    compat_local_recording_metadata,
    compat_func=_local_recording_metadata,
    recording_metadata_map=recording_metadata_map,
    format_duration=format_duration,
)
cloud_recording_metadata = partial(
    compat_cloud_recording_metadata,
    compat_func=_cloud_recording_metadata,
    recording_metadata_map=recording_metadata_map,
)
remember_drive_file = partial(
    compat_remember_drive_file,
    compat_func=_remember_drive_file,
    load_json_data=load_json_data,
    save_json_data=save_json_data,
)
forget_drive_file = partial(
    compat_forget_drive_file,
    compat_func=_forget_drive_file,
    load_json_data=load_json_data,
    save_json_data=save_json_data,
)
def sheet_payload():
    return _sheet_payload(load_json_data("sheet_library"))


recording_file_bytes = partial(compat_recording_file_bytes, compat_func=_recording_file_bytes)
convert_path_to_mp3 = partial(_convert_path_to_mp3, logger=logger)
get_audio_duration_seconds = partial(_get_audio_duration_seconds, logger=logger)
remember_recording_duration = partial(
    compat_remember_recording_duration,
    compat_func=_remember_recording_duration,
    load_json_data=load_json_data,
    save_json_data=save_json_data,
    next_id=next_id,
    format_duration=format_duration,
)

# Root and health endpoints moved to src/backend/routers/meta.py.
# Maintenance orphan endpoint moved to src/backend/routers/maintenance.py.
# System database endpoints were moved to src/backend/routers/system.py.
# Basic CRUD endpoints live in src/backend/routers/*.py.
# Endpoints moved to src/backend/routers/recordings.py.
# Endpoints moved to src/backend/routers/scores.py.


recording_payload = partial(compat_recording_payload, compat_func=_recording_payload, load_json_data=load_json_data, format_duration=format_duration)
local_recording_path = partial(compat_local_recording_path, compat_func=_local_recording_path)


def normalize_extra_payload(payload: dict[str, Any], item_id: int | None = None, current: dict[str, Any] | None = None) -> dict[str, Any]:
    from .services.extra_collection_helpers import normalize_extra_payload as _normalize_extra_payload

    return _normalize_extra_payload(payload, next_updated_at_func=next_updated_at, item_id=item_id, current=current)


def collection_items(name: str) -> list[dict[str, Any]]:
    from .services.extra_collection_helpers import collection_items as _collection_items

    return _collection_items(name, load_json_data)

# Extra and album endpoints moved to src/backend/routers/albums.py.
# System administration endpoints were moved to src/backend/routers/system.py.


__all__ = sorted(APP_CORE_EXPORTS)


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app, host="0.0.0.0", port=int(os.getenv("PORT", "8080")))
