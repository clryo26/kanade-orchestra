from __future__ import annotations

import logging
from typing import Any

from fastapi import HTTPException, UploadFile
try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional dependency guard
    load_workbook = None

from ..core.runtime_paths import TIMETABLE_TEMPLATE_PATH
from ..repositories.performance_repository import PerformanceRepository
from .image_asset_service import delete_stored_image, ensure_public_image_url, is_data_image, store_data_image, store_uploaded_image
from .storage_service import load_json_data
from .timetable_payload_helpers import build_timetable_workbook_bytes, excel_safe_filename, performance_day_info_for_performance

_repo = PerformanceRepository()
logger = logging.getLogger(__name__)


def _flyer_route(performance_id: int) -> str:
    return f"/api/performances/{performance_id}/flyer-image"


def _flyer_object_prefix(performance_id: int) -> str:
    return f"performance-flyers/{performance_id}/flyer"


async def _persist_flyer_image(
    performance_id: int,
    flyer_image: str,
    flyer_file: UploadFile | None = None,
) -> str:
    if flyer_file is not None:
        return await store_uploaded_image(
            flyer_file,
            object_prefix=_flyer_object_prefix(performance_id),
            route_path=_flyer_route(performance_id),
        )
    return store_data_image(
        flyer_image,
        object_prefix=_flyer_object_prefix(performance_id),
        route_path=_flyer_route(performance_id),
    )


def _public_flyer_image(performance_id: int, flyer_image: str) -> str:
    return ensure_public_image_url(
        flyer_image,
        route_path=_flyer_route(performance_id),
    )


def _delete_flyer_image(performance_id: int, flyer_image: str) -> None:
    delete_stored_image(flyer_image, object_prefix=_flyer_object_prefix(performance_id))


def _safe_fee_amount(value: Any) -> float:
    if value in (None, ""):
        return 0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0


def _normalize_piece(piece: Any) -> Any | None:
    if isinstance(piece, dict):
        title = str(piece.get("title") or piece.get("piece") or piece.get("name") or "").strip()
        if not title:
            return None
        normalized = dict(piece)
        normalized["title"] = title
        return normalized
    title = str(piece or "").strip()
    if not title:
        return None
    return title


def _normalize_performance(payload: dict[str, Any]) -> dict[str, Any]:
    normalized = dict(payload)
    for key in ("title", "date", "open_time", "start_time", "venue", "conductor", "flyer_image"):
        normalized[key] = "" if normalized.get(key) is None else str(normalized.get(key, ""))
    normalized["performance_fee_amount"] = _safe_fee_amount(normalized.get("performance_fee_amount"))

    raw_pieces = normalized.get("pieces")
    pieces = raw_pieces if isinstance(raw_pieces, list) else []
    normalized["pieces"] = [
        piece
        for raw_piece in pieces
        if (piece := _normalize_piece(raw_piece)) is not None
    ]
    return normalized


def _save_error(operation: str, exc: Exception) -> HTTPException:
    logger.exception("Performance %s failed", operation)
    return HTTPException(status_code=500, detail=f"Performance {operation} failed: {exc}")


def list_performances() -> list[dict[str, Any]]:
    items = [_normalize_performance(item) for item in _repo.list_all()]
    for item in items:
        performance_id = int(item.get("id") or 0)
        if performance_id:
            item["flyer_image"] = _public_flyer_image(performance_id, item.get("flyer_image") or "")
    return items


def get_performance(performance_id: int) -> dict[str, Any]:
    _, item = _repo.find_by_id(performance_id)
    normalized = _normalize_performance(item)
    normalized["flyer_image"] = _public_flyer_image(performance_id, normalized.get("flyer_image") or "")
    return normalized


def get_performance_record(performance_id: int) -> dict[str, Any]:
    _, item = _repo.find_by_id(performance_id)
    return item


async def create_performance(payload: dict[str, Any], flyer_file: UploadFile | None = None) -> dict[str, Any]:
    created: dict[str, Any] | None = None
    try:
        normalized_payload = _normalize_performance(payload)
        flyer_source = str(normalized_payload.get("flyer_image") or "")
        if flyer_file is not None or is_data_image(flyer_source):
            normalized_payload["flyer_image"] = ""
        created = _normalize_performance(_repo.create(normalized_payload))
        performance_id = int(created.get("id") or 0)
        flyer_image = await _persist_flyer_image(performance_id, flyer_source, flyer_file)
        created["flyer_image"] = flyer_image
        if flyer_image != str(normalized_payload.get("flyer_image") or ""):
            _repo.update(performance_id, lambda current: {**current, "flyer_image": flyer_image})
        return created
    except HTTPException:
        if created is not None:
            performance_id = int(created.get("id") or 0)
            flyer_image = str(created.get("flyer_image") or "")
            if flyer_image:
                _delete_flyer_image(performance_id, flyer_image)
            _repo.delete(performance_id)
        raise
    except Exception as exc:
        if created is not None:
            performance_id = int(created.get("id") or 0)
            flyer_image = str(created.get("flyer_image") or "")
            if flyer_image:
                _delete_flyer_image(performance_id, flyer_image)
            _repo.delete(performance_id)
        raise _save_error("create", exc) from exc


async def update_performance(
    performance_id: int,
    payload: dict[str, Any],
    flyer_file: UploadFile | None = None,
) -> dict[str, Any]:
    normalized_payload = _normalize_performance(payload)
    normalized_payload.pop("created_at", None)
    uploaded_flyer_url = ""
    try:
        _, current = _repo.find_by_id(performance_id)
        old_flyer_image = str(current.get("flyer_image") or "")
        merged = _normalize_performance({**current, **normalized_payload})
        flyer_source = str(merged.get("flyer_image") or "")
        uploaded_flyer_url = await _persist_flyer_image(performance_id, flyer_source, flyer_file)
        merged["flyer_image"] = uploaded_flyer_url
        updated = _normalize_performance(_repo.update(performance_id, lambda _current: merged))
        if old_flyer_image and old_flyer_image != str(updated.get("flyer_image") or ""):
            _delete_flyer_image(performance_id, old_flyer_image)
        return updated
    except HTTPException:
        if uploaded_flyer_url:
            _delete_flyer_image(performance_id, uploaded_flyer_url)
        raise
    except Exception as exc:
        if uploaded_flyer_url:
            _delete_flyer_image(performance_id, uploaded_flyer_url)
        raise _save_error("update", exc) from exc


def delete_performance(performance_id: int) -> None:
    _, current = _repo.find_by_id(performance_id)
    _delete_flyer_image(performance_id, str(current.get("flyer_image") or ""))
    _repo.delete(performance_id)


def build_timetable_report(performance_id: int) -> tuple[bytes, str]:
    performance = get_performance(performance_id)
    info = performance_day_info_for_performance(performance_id, load_json_data)
    if not info:
        raise HTTPException(status_code=404, detail="performance_day_info not found")

    workbook_bytes = build_timetable_workbook_bytes(
        performance,
        info,
        load_workbook_func=load_workbook,
        template_path=TIMETABLE_TEMPLATE_PATH,
    )
    date_text = str(performance.get("date") or "").strip()
    title_text = excel_safe_filename(performance.get("title") or "performance")
    filename = (
        f"{date_text}_{title_text}_本番タイムテーブル.xlsx"
        if date_text
        else f"{title_text}_本番タイムテーブル.xlsx"
    )
    return workbook_bytes, filename
